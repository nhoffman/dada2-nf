#!/usr/bin/env python3
import csv
import itertools
import sys
import argparse
from os import path
from glob import glob
import json
import openpyxl
import operator
import os
import subprocess


def read_manifest_excel(fname):
    """Read the first worksheet from an excel file and return a generator
    of dicts with keys limited to 'keepcols'.

    """

    valgetter = operator.attrgetter('value')

    wb = openpyxl.load_workbook(fname)
    sheet = wb[wb.sheetnames[0]]

    rows = (r for r in sheet.iter_rows() if r[0].value)

    # get fieldnames from cells in the first row up to the first empty one
    header = itertools.takewhile(valgetter, next(rows))
    fieldnames = [cell.value for cell in header]
    yield fieldnames
    for row in rows:
        d = dict(zip(fieldnames, map(valgetter, row)))
        if d['sampleid']:
            yield d


def main(arguments):

    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('plate', help="plate label, ie miseq-plate-{label}")
    parser.add_argument('-d', '--data-dir',
                        default='/fh/fast/fredricks_d/bvdiversity/data')
    parser.add_argument('-o', '--out-dir', default='dada2_nf_out')
    parser.add_argument('-f', '--force', action='store_true', default=False,
                        help="""do not prompt for confirmation before
                        launching pipleine""")
    parser.add_argument('-c', '--check-inputs', action='store_true', default=False,
                        help='verify inputs and exit')
    parser.add_argument('-p', '--profile', choices=['hutch'],
                        default='hutch_singularity', help="""[%(default)s]""")

    args = parser.parse_args(arguments)
    plate = args.plate

    platedir = path.join(args.data_dir, f'miseq-plate-{plate}')
    outdir = path.join(args.data_dir, args.out_dir, f'miseq-plate-{plate}')
    try:
        os.makedirs(outdir)
    except OSError as err:
        pass

    # fastq list - the directory layout has changed over time, so we
    # need to try a few patterns
    layouts = [
        f'run-files/*/Data/Intensities/BaseCalls/m{plate}*.fastq.gz',
        f'run-files/*/Alignment_1/*/Fastq/m{plate}*.fastq.gz',  # most recent format
        f'run-files/*/*/m{plate}*.fastq.gz',  # eg, plate 3
        f'run-files/*/*/*/m{plate}*.fastq.gz',  # plate 91
    ]

    for pattern in layouts:
        fq_pattern = path.join(platedir, pattern)
        print(f'searching {fq_pattern}')
        files = glob(fq_pattern)
        print('found {} fastq files'.format(len(files)))
        if files:
            break
    else:
        print('no files found, exiting')
        sys.exit(1)

    fastq_list = path.abspath(path.join(outdir, 'fastq_list.txt'))
    with open(fastq_list, 'w') as f:
        f.write('\n'.join(files) + '\n')

    # sample information
    sample_information = path.join(
        platedir, f'sample-information/sample-information-m{plate}.xlsx')
    assert os.path.exists(sample_information), f'{sample_information} not found'
    manifest = read_manifest_excel(sample_information)
    fieldnames = next(manifest)
    writer = csv.DictWriter(
        open(path.join(outdir, 'sample_information.csv'), 'w'),
        fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(manifest)

    params_file = path.join(outdir, 'params.json')
    d = {
        'sample_information': sample_information,
        'fastq_list': fastq_list,
        'output': path.join(outdir, 'output'),
        'index_file_type': 'dual',
        'dada_params': 'data/dada_params_300.json',
        'alignment': {
            'library': "",
            'model': 'data/SSU_rRNA_bacteria.cm',
            'strategy': 'cmsearch'
            }
    }
    with open(params_file, 'w') as f:
        json.dump(d, f, indent=2)


    if args.check_inputs:
        return

    version_info = path.join(outdir, 'version_info.txt')
    cmds = [
        f'date > {version_info}',
        f'pwd >> {version_info}',
        f'git describe --tags --dirty >> {version_info}',
        f'git log -n1 >> {version_info}',
    ]
    for cmd in cmds:
        subprocess.run(cmd, shell=True)

    runcmd = (f'nextflow run main.nf '
              f'-profile {args.profile} '
              f'-params-file {params_file}')

    if args.profile == 'singularity':
        runcmd += f' --work_dir workdirs/plate-{plate}'

    print(runcmd)
    response = 'yes' if args.force else input('Run this pipleine? (yes/no) ')
    if response == 'yes':
        p = subprocess.run(
            runcmd, shell=True,
            env=dict(os.environ, NXF_SINGULARITY_CACHEDIR='singularity_cache'))
        return p.returncode

if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
