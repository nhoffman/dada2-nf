#!/usr/bin/env python3

"""
Required fields for the manifest:
* sampleid - a string found in the fastq file name uniquely identifying a
  specimen
* batch - a label grouping specimens into PCR batches for dada2::learnErrors()

Required fields if no fastq_list in params:
* R1 - Path to the R1 FASTQ file
* R2 - Path to the R2 FASTQ file
* I1 - Path to the I1 index FASTQ file
* I2 - Path to the I2 index FASTQ file

Verifies the following:
* all sampleids are unique
* every sampleid in the manifest has corresponding R{1,2} and I{1,2}

"""

import argparse
import csv
import gzip
import itertools
import operator
import os
import re
import sys

import openpyxl


# FASTQ sample IDs are parsed in two steps:
# 1. READ_SUFFIX removes the read/index marker and everything after it.
# 2. SAMPLE_SUFFIX removes optional Illumina sample/lane suffixes.
#
# Examples:
#   m3n701-s502_S1_L001_R1_001.fastq.gz -> m3n701-s502
#   22R255-NGSITS49_S4_L001_I2_001.fastq.gz -> 22R255-NGSITS49
#   1_07_5479_wk12_R2_001.fastq.gz -> 1_07_5479_wk12
SAMPLE_SUFFIX = re.compile(r'_S\d+(_L\d{3})?$')
READ_SUFFIX = re.compile(r'_[RI][12]_.*$')
READ_ORDER = {'I1': 0, 'I2': 1, 'R1': 2, 'R2': 3}


def read_manifest_excel(fname):
    """Read the first worksheet from an excel file and return a generator
    of dicts with keys limited to 'keepcols'.

    """

    valgetter = operator.attrgetter('value')

    wb = openpyxl.load_workbook(fname)
    sheet = wb[wb.sheetnames[0]]

    rows = (r for r in sheet.iter_rows() if r[0].value)

    # get fieldnames from cells in the first row up to the first empty one
    header = itertools.takewhile(valgetter, next(rows))

    # replace column names to be discarded with '_'
    fieldnames = [cell.value for cell in header]

    for row in rows:
        d = dict(zip(fieldnames, map(valgetter, row)))
        if d['sampleid']:
            yield d


def read_manifest_csv(fname):

    with open(fname) as f:
        f = (li for li in f if not li.strip().startswith('#'))
        reader = csv.DictReader(f)
        for d in reader:
            if d['sampleid']:
                yield dict(d)


def get_sampleid(pth):
    sampleid = READ_SUFFIX.sub('', os.path.basename(pth))
    return SAMPLE_SUFFIX.sub('', sampleid)


def get_read_order(pth):
    basename = os.path.basename(pth)
    for read, order in READ_ORDER.items():
        if '_' + read + '_' in basename:
            return order
    raise ValueError('unknown read in filename: ' + pth)


def main(arguments):
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('manifest')
    parser.add_argument('fastq_list', type=argparse.FileType('r'))
    parser.add_argument('--index-file-type',
                        choices=['single', 'dual', 'none'],
                        default='dual',
                        help='dual, single, or no index files?')
    outputs = parser.add_argument_group('outputs')
    outputs.add_argument('--batches',
                         help="Output csv file mapping sampleid to batch",
                         type=argparse.FileType('w'))
    outputs.add_argument('--counts',
                         help="raw fastq_file counts",
                         type=argparse.FileType('w'))
    outputs.add_argument('--sample-index',
                         help=("Output csv file mapping index "
                               "files to sampleid and R1/R2 files"),
                         type=argparse.FileType('w'))
    args = parser.parse_args(arguments)

    # sort fqs by natural alpha order [_I1_, _I2_, _R1_, _R2_]
    fq_files = (f.strip() for f in args.fastq_list)
    fq_files = sorted(
        (f for f in fq_files if f and not f.startswith('#')),
        key=lambda f: (get_sampleid(f), get_read_order(f), f))

    if args.manifest.endswith('.csv'):
        read_manifest = read_manifest_csv
    else:
        read_manifest = read_manifest_excel

    manifest_reader = read_manifest(args.manifest)
    manifest_data = list(manifest_reader)
    manifest = {m['sampleid']: m for m in manifest_data}

    labels = {
        'dual': ['I1', 'I2', 'R1', 'R2'],
        'single': ['I1', 'R1', 'R2'],
        'none': ['R1', 'R2'],
    }

    expected = labels[args.index_file_type]
    unexpected = set(labels['dual']) - set(expected)

    # ## some basic manifest sanity checks ###
    # make sure all sampleids are unique
    assert len([m['sampleid'] for m in manifest_data]) == len(manifest)

    fq_sampleids = {get_sampleid(pth) for pth in fq_files}

    # confirm that all sampleids in the manifest have corresponding fastqs
    extras = manifest.keys() - fq_sampleids
    if extras:
        sys.exit('samples in the manifest without fastq files: {}'.format(
            extras))

    # confirm that all fastq files are in manifest
    extras = fq_sampleids - manifest.keys()
    if extras:
        sys.exit('fastq not present in manifest: {}'.format(extras))

    # confirm that every sampleid has required fastq files
    for sampleid, fns in itertools.groupby(fq_files, key=get_sampleid):
        labels = [e for fi in fns for e in expected if f'_{e}_' in fi]
        if labels != expected:
            sys.exit('a fastq file is missing for sampleid {}: has {}'.format(
                sampleid, labels))

    # warn if extra fastqs are present per sampleid
    for sampleid, fns in itertools.groupby(fq_files, key=get_sampleid):
        extras = [fi for fi in fns for e in unexpected if f'_{e}_' in fi]
        if extras:
            print('WARNING: extra fastq file(s) '
                  'present for sampleid {}: {}'.format(sampleid, extras))
    # ###

    for i in manifest:
        if not manifest[i]['batch']:
            manifest[i]['batch'] = 'unknown'

    fq_files = [fi for fi in fq_files for e in expected if f'_{e}_' in fi]

    # remove expected manifest columns
    for i in manifest:
        for e in unexpected:
            if e in manifest[i]:
                del manifest[i][e]

    # set manifest fq paths
    for sampleid, fqs in itertools.groupby(fq_files, key=get_sampleid):
        manifest[sampleid].update(zip(expected, fqs))

    # outputs
    manifest = manifest.values()

    if args.sample_index:
        out = csv.DictWriter(
            args.sample_index,
            fieldnames=['sampleid', 'direction', 'fastq', 'I1', 'I2'],
            extrasaction='ignore')
        for m in manifest:
            for d in ['R1', 'R2']:
                out.writerow({'direction': d, 'fastq': m[d], **m})

    if args.batches:
        out = csv.DictWriter(
            args.batches,
            fieldnames=['sampleid', 'batch'],
            extrasaction='ignore')
        out.writerows(manifest)

    if args.counts:
        out = csv.DictWriter(args.counts, fieldnames=['sampleid', 'count'])
        out.writeheader()
        for m in manifest:
            fq = gzip.open(os.path.basename(m.get('I1', None) or m['R1']))
            count = sum(1 for li in fq if li.startswith(b'+'))
            out.writerow({'sampleid': m['sampleid'], 'count': count})


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
